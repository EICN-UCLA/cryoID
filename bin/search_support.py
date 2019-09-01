#!/usr/bin/env python2
# -*- coding: utf-8 -*-

"""
search_support.py contains several supporting functional modules for search_pool.py, such as RetrievingMapping/DBGenerate/SeqTrans/SeqBlast/SeqSelect.

Author: Xiaorun Li (Lee) @ Prof. Hong Zhou lab
University of California, Los Angeles (UCLA) &
University of Science and Technology of China (USTC)

Version 1.0, released in 9/1/2019
"""

import os, sys, subprocess, logging
from argparse import Namespace

#=======================================================================================
# RetrevingMapping module
# To get protein candidate sequences from MS results (.xlsx) or Genename list (.txt) 
# Code modified from http://www.uniprot.org/help/api_idmapping#id_mapping_python_example
#=======================================================================================
import urllib,urllib2 #, ssl

# Retrieve/ID mapping website
retrieve_url = 'https://www.uniprot.org/uploadlists/'

# connect to uniprot.org and do the retrieving job automatically. return the fasta/info files
# params.candidate_pool = candidate_pool_name (a file containing candidate protein names)
def RetrievSeq(params):
    
    # logging setup
    logging.basicConfig(filename=params.logfile, level=logging.INFO, format='%(message)s')

    # get the filename
    file_name = os.path.basename(params.candidate_pool)
    file_base, file_exten = os.path.splitext(file_name)
    fasta_outfile = file_base + '.fasta'
    #info_outfile = file_base + '_info.txt'
        
    # check if outputfiles already exist
    if os.path.exists(fasta_outfile): # and os.path.exists(info_outfile):
        #print 'Previous sequence pool files found. Use them for current run.'
        # print output immediately
        #sys.stdout.flush()
        pass
    else:
        
        print 'Retreving/ID mapping from Uniprot...\n'
        # print output immediately
        sys.stdout.flush()
        logging.info('Retreving/ID mapping from Uniprot...\n')
        
        # read all the gene name			
        proteins = []
        
        '''
        if file_name.endswith('.xlsx'):
            # import openpyxl for reading and writing .xlsx file
            from openpyxl import load_workbook # Workbook        
                
            # load the .xlsx file and read the sheet	
            wb = load_workbook(params.candidate_pool)	
            sheet = wb[wb.sheetnames[0]]
            	
            # read the header and find the column of gene name
            header = []	
            for col in sheet.iter_cols():
                header.append(col[0].value)
            	
            try:	
                target = header.index('Protein')
            except ValueError:
                logging.exception('Cannot find Protein list in the sequence pool file. Exiting...')
                raise ValueError('Cannot find Protein list in the sequence pool file. Exiting...')
            			
            for row in sheet.iter_rows(row_offset=1):
                proteins.append(row[target].value)		
                
            proteins.pop()   
        
        # file_name ends with .txt
        elif file_name.endswith('.txt') or file_name.endswith('.list'):
        '''
        # check if file name ends with .txt/.list
        if file_name.endswith('.txt') or file_name.endswith('.list'):
            f0 = open(params.candidate_pool, 'r')
            
            for protein in f0:
                if protein != '\n' and protein[0] != '#':
                    proteins.append(protein[:-1])
            
            f0.close()


        # the following code is modified from www.uniprot.org
        '''
        info_retrieve = {
        #'from': 'GENENAME',
        'from': 'ACC+ID',
         # 'from': 'ACC',
        'to': 'ACC',
        'format': 'tab',
        'query': ' '.join(str(e) for e in proteins),
        'columns': 'id%2Cgenes%2Clength'
        }
        '''
        
        fasta_retrieve = {
        #'from': 'GENENAME',
        'from': 'ACC+ID',
         # 'from': 'ACC',
        'to': 'ACC',
        'format': 'FASTA',
        'query': ' '.join(str(e) for e in proteins)       	
        }
        	        

        # get the mapping results. Codes modified from https://www.uniprot.org/help/api_idmapping
        try:
            
            # to get rid of urllib and “SSL: CERTIFICATE_VERIFY_FAILED” Error
            # https://stackoverflow.com/questions/27835619/urllib-and-ssl-certificate-verify-failed-error
            #gcontext = ssl.SSLContext(ssl.PROTOCOL_TLSv1)
            '''
            info_data = urllib.urlencode(info_retrieve)
            info_request = urllib2.Request(retrieve_url, info_data)
            info_response = urllib2.urlopen(info_request) #, context=gcontext)
            
            info_lines = info_response.read()
                        
            if info_lines:
                # write the database file  
                f1 = open(info_outfile, "w")
            
                for line in info_lines.split('\n'):
                    if line:
                        line = line.split('\t')[0:3]
                        f1.write(line[0] + '\t' + line[1] + '\t' + line[2] + '\n')          
                #f1.write(tab_response.read()) 
                f1.close()
            else:
                logging.exception('')
                raise RuntimeError
            '''    
            
            fasta_data = urllib.urlencode(fasta_retrieve)
            fasta_request = urllib2.Request(retrieve_url, fasta_data)            
            fasta_response = urllib2.urlopen(fasta_request) #, context=gcontext)
            
            fasta_lines = fasta_response.read()
            
            if fasta_lines:
                f2 = open(fasta_outfile, "w")    
                f2.write(fasta_lines)
                f2.close()
                
            else:
                raise RuntimeError
                    
        except RuntimeError:
            logging.exception('Error when retrieving/mapping the candate protein sequences!')
            raise RuntimeError('Error when retrieving/mapping the candate protein sequences!')

    # return the output filenames
    #return [fasta_outfile, info_outfile]
    return fasta_outfile


#=======================================================================================

# Sequence grouping/mapping tools
# Map sequences for blastp searching or group amino acids to show user the degenerate sequences
#=======================================================================================

from collections import deque

# Amino acid grouping and mapping table for blastp search. Representive codes are optimised based on PAM30 scorig matrix.
# include lower case codes and rare amino acids 
aaMapping = {
    'G':'G', 'A':'G', 'S':'G', 'C':'G', 'V':'G', 'T':'G', 'I':'G',
    'g':'G', 'a':'G', 's':'G', 'c':'G', 'v':'G', 't':'G', 'i':'G',    # lower case
    'P':'P', 'p':'P',
    'L':'Z', 'D':'Z', 'N':'Z', 'E':'Z', 'Q':'Z', 'M':'Z', #'I':'G',
    'l':'Z', 'd':'Z', 'n':'Z', 'e':'Z', 'q':'Z', 'm':'Z', #'i':'G', 
    'K':'M', 'R':'M',
    'k':'M', 'r':'M',
    #'K':'Y', 'R':'Y', 'k':'Y', 'r':'Y',
    'H':'Y', 'F':'Y', 'Y':'Y',
    'h':'Y', 'f':'Y', 'y':'Y',
    'W':'W', 'w':'W',
    'X':'X', 'x':'X',
    
     # Rare amino acid
    'U':'G', 'u':'G',         # Selenocysteine is a cysteine analogue with a selenium-containing selenol group in place of the sulfur-containing thiol group.
    'B':'Z', 'Z':'Z',         # asparagine/aspartic acid - asx - B, glutamine/glutamic acid - glx - Z
    'b':'Z', 'z':'Z',
    #'O':'M', 'o':'M'          # Pyrrolysine pyrroline side-chain is similar to that of lysine in being basic and positively charged at neutral pH. 
}

# Reverse Mapping table from the blastp codes to the 6 simplified groups
aaRevMapping = {
    'G':'G', 'P':'P', 'Z':'L','M':'K','Y':'Y', 'W':'W', 'X':'X'
}


# reference query table/dictionary based on the simulation. assume 10% tolerance error.
# if query length is shorter than specified, print warning infomation
# query_num: cutoff_query_length. O for query_num > 6
query_set = {1: 30, 2: 20, 3: 16, 4: 15, 5: 13, 6: 12, 0: 12}

# Group the amino acids and yield degenearete sequences for user inspection
# params.seqfile=queryfile (query file name)
def SeqGroup(params):
    
    # logging setup
    logging.basicConfig(filename=params.logfile, level=logging.INFO, format='%(message)s')
    
    f1 = open(params.seqfile,'r')
    outfile = params.seqfile[:-6] + '_simplified.fasta'
    o1 = open(outfile, 'w')
    
    # count the sequence number/length
    seq_num = 0
    seq_len = 0
    
    for seq in f1:
        if seq and seq[0][0] != '>' and seq[0][0] != '#':            
            try:
                seq = ''.join(aaMapping[e] for e in seq[:-1])
                seq = ''.join(aaRevMapping[e] for e in seq)
                
                # count the query number/length
                seq_num += 1
                seq_len += len(seq)
            except KeyError as error:
                logging.exception('Unrecognized amino acid ' + str(error) + ' in line ' + seq[:-1] + ' of ' + params.seqfile)
                raise KeyError('Unrecognized amino acid ' + str(error) + ' in line ' + seq[:-1] + ' of ' + params.seqfile)
                
            seq += '\n'
            
        o1.write('%s' %(seq))

    f1.close()
    o1.close()
    
    #return seq_num, seq_len, outfile

    
# Map the natural sequences to degenerate codes for blastp searching
# params.seqfile = queryfile (query file name)
# params.reverse = True/False (whether include the reverse sequence)
def SeqMap(params):

    # logging setup
    logging.basicConfig(filename=params.logfile, level=logging.INFO, format='%(message)s')
    
    f1 = open(params.seqfile, 'r')
    outfile = params.seqfile[:-6] + '_mapped.fasta'
    o1 = open(outfile, 'w')
            
    # count total number of amino acid in input sequence
    icount = 0
    
    # use queue to store and then pop lines. More efficient than list
    tem = deque([])	

    for seq in f1:
        if seq and seq[0][0] != '>' and seq[0][0] != '#':
            #for k, v in aaMapping.iteritems():
                #seq = seq.replace(k, v)
            # replace the codes accroding to the table
            try:
                seq = ''.join(aaMapping[e] for e in seq[:-1])
            except KeyError as error:
                logging.exception('Unrecognized amino acid ' + str(error) + ' in line ' + seq[:-1] + ' of ' + params.seqfile)
                raise KeyError('Unrecognized amino acid ' + str(error) + ' in line ' + seq[:-1] + ' of ' + params.seqfile)
                
            seq += '\n'
                
            icount += len(seq) - 1
             
        o1.write('%s' %(seq))	
		
        # to inverse the sequence and write it to the file
        # watch out for '\n' at the end of each line 
        if params.reverse:
            if seq[0] != '>':
                seq =seq[:-1]
                seq = seq[::-1] + '\n'
            else:
                seq = seq[:-1] + '_rev\n'

            tem.append(seq)

    while tem:
        o1.write('%s' %(tem.popleft()))
	
    f1.close()
    o1.close()
	
    return icount, outfile

#===============================================================================================

# generate the (degenerate) query sequence files for blastp searching. One query, one file.
#===============================================================================================
# params.queryfile = queryfile (query file name)
# params.reverse = True/False (whether include the reverse sequence)
# params.verbose = True/False (whether show verbose results)
def QueryMap(params):
    
    # logging setup
    logging.basicConfig(filename=params.logfile, level=logging.INFO, format='%(message)s')
    
    # if pdb file is provided, print the sequences from the pdb model with phenix.print_sequence. Note that mse represents X in coot.
    if params.queryfile.endswith('.pdb'):
        print_seq = ['phenix.print_sequence', params.queryfile, '--letter_for_mse=X']
        
        # call phenix.print_sequence as subprocess and get the output/error message
        try:
            proc = subprocess.Popen(print_seq, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            (out, error) = proc.communicate()
        except OSError:
            logging.exception('Could not find command phenix.print_sequence. Make sure you include it in your PATH!')
            raise OSError('Could not find command phenix.print_sequence. Make sure you include it in your PATH!')
            
        if error:
            logging.exception('Error when calling phenix.print_sequence to convert pdb file to fasta file\n' + error)
            raise OSError('Error when calling phenix.print_sequence to convert pdb file to fasta file\n' + error)
            
        # create the .fasta file
        queryfile = params.queryfile.split('.pdb')[0] + '.fasta'
        f0 = open(queryfile, 'w')
        print >> f0, out
        
        f0.close()
        
    # if fasta file provided, directly use it
    elif params.queryfile.endswith('.fasta'):
        queryfile = params.queryfile
    
    # generate the degenerate query sequence file for the user    
    if params.verbose:
        group_params = Namespace(seqfile=queryfile, logfile=params.logfile)
        #query_num, query_len, query_grouped = SeqGroup(group_params)
        SeqGroup(group_params)
    
    
    # split queries. One query, one file.
    f1 = open(queryfile, 'r')
    
    # count query number
    iquery = 1
    # count total length of queries
    query_lens = 0
    # count residue X number
    x_num = 0
    
    # lists to save query length and query file names
    query_info = []
    
    # loop over all queries
    while True:
        
        # read two lines every time, query name and sequence (one-to-one). Notice that there may be comment line (start with '#')
        # read description line first
        line1 = f1.readline()
        if not line1: 
            break
        if line1[0] != '>':
            continue
        # read sequence line        
        line2 = f1.readline()
                        
        #if len(line2) < 12 or len(line2) > 40:
            #raise ValueError('Query '+ line1[1:-1] + ' length is beyond the workable range 12~40. Exiting')
                            
        # Assign each query a number in case they share a same name. remove any separate parameters
        query_name = ''.join(line1[1:].split()) + '_' + str(iquery)
        query_split = queryfile[:-6] + '_' + str(iquery) + '.fasta'
        
        # write the current query to file
        o1 = open(query_split, 'w')    
        o1.write('>' + query_name + '\n')
        o1.write(line2)     
        o1.close()
        
        # tranform query and return length
        map_params = Namespace(seqfile=query_split, reverse=params.reverse, logfile=params.logfile)
        query_len, query_file = SeqMap(map_params)
        
        # add current query length
        query_lens += query_len
        
        # count X number in current query sequence
        x_num_curr = 0
        for c in line2:
            if c == 'X' or c == 'x':
                x_num_curr += 1
        x_num += x_num_curr
        
        # remove the intermediate file
        os.remove(query_split)
        
        # add the mapped query name, length, file name and number of X to the query_info list
        query_info.append([query_name, query_len, query_file, x_num_curr])
        
        iquery += 1
    
    # get the query number
    query_num = iquery - 1    
    
    if not query_num:
        logging.exception('No query found. Check your input! Exiting')
        raise OSError('No query found. Check your input! Exiting')
    elif query_num > 10:
        logging.warning('The number of queries are more than 10! Make sure you selected the right query file!')
        print 'Warning: The number of queries are more than 10! Make sure you selected the right query file!'
        sys.stdout.flush()
    elif query_num > 50:
        logging.exception('The number of queries are more than 50! Check your input query file again! Exiting...')
        raise OSError('The number of queries are more than 50! Check your input query file again! Exiting...')
    
    
    # calculate total predicted animo acids in the query (not including residue X)
    aa_num = query_lens - x_num
    # calculate the average query length
    ave_aa_num = int(round(aa_num * 1.0 /query_num))
    
    # get the average X residues in the quary sequences
    ave_x_num = x_num * 1.0 / query_num
    #x_num_percent = x_num * 1.0 / query_lens

    # print the query set info
    #print '\n####################################################################'
    print '********************************************************************'
    print 'QUERY SEQUENCE SUMMARY\n\tNumber of query sequences used:  '+str(query_num)+'\n\tNumber of identifiable residues per query:  '+str(ave_aa_num)+\
    '\n\tNumber of unidentifiable residues per query:  '+'{0:.2g}'.format(ave_x_num) #+' ({0:.1%})'.format(x_num_percent) 
    sys.stdout.flush()

    # print these info to log file too
    logging.info('QUERY SEQUENCE SUMMARY\n\tNumber of query sequences used:  '+str(query_num)+'\n\tNumber of identifiable residues per query:  '+str(ave_aa_num)+\
    '\n\tNumber of unidentifiable residues per query:  '+'{0:.2g}'.format(ave_x_num)) #+' ({0:.1%})'.format(x_num_percent)
    
    
    # check if the query is too short based on the simulation results (10% tolerance error)
    if query_num > 6:
        cutoff = query_set[0]
    else:
        cutoff = query_set[query_num]
    
    if ave_aa_num < cutoff:
        print 'Warning: query length is short! You may consider extending your queries, or the final results may be not reliable!'
        sys.stdout.flush()
        logging.warn('Warning: query length is short! You may consider extending your queries, or the final results may be not reliable!')
        
    print ''
    sys.stdout.flush()
    logging.info('')
    
    return query_info


#===============================================================================================

# Map candidate sequences and make local blastp database by calling makeblastdb in blast package
#===============================================================================================

# default parameters for makeblastdb
dbtype = 'prot'            # molecular type of target db, default protein   

# generate local blastp database with (degenerate) sequences and get database size
# params.seqfile = seqfile (sequence pool file name, fasta format)
def DBGenerate(params):
    
    # logging setup
    logging.basicConfig(filename=params.logfile, level=logging.INFO, format='%(message)s')
    
    # get database name
    if params.seqfile.endswith('.fasta'):
        dbname = './database/' + params.seqfile[:-6] + '_db'
    else:
        logging.exception('Seqfile '+ params.seqfile +' is not of .fasta format! Exiting')
        raise TypeError('Seqfile '+ params.seqfile +' is not of .fasta format! Exiting')
        
    # by default the database is located in ./database subfolder
    if os.path.exists(dbname + '.pin'):
        #print 'Database found. Use the previous one for the current searching.'
        # print output immediately
        #sys.stdout.flush()
        
        # check txt file to get database size or count the database size from seqfile
        if os.path.exists(dbname + '.txt'):
            f1 = open(dbname + '.txt', 'r')
            line = f1.readline()
            try:
                dbsize = int(line.split('\t')[1])
            except TypeError:
                logging.exception('Error when getting database size from' + dbname + '.txt. Check the file!')
                raise TypeError('Error when getting database size from' + dbname + '.txt. Check the file!')
                
            f1.close()
            
        else:    
            dbsize = 0
            f2 = open(params.seqfile,'r')
            for seq in f2:
                if seq[0][0] != '>':
                    dbsize += len(seq) - 1
                                        
            f2.close()
                
    else:
        #print 'Generating blastp local database...'
        # print output immediately
        sys.stdout.flush()
        # map the sequences and get database size
        map_params = Namespace(seqfile=params.seqfile, reverse=False, logfile=params.logfile)
        dbsize, mapped_seqfile = SeqMap(map_params)
	
        # run makeblastdb to build the local database
        mkdb_arg = ['makeblastdb', '-in', mapped_seqfile, '-out', dbname, '-dbtype', dbtype, '-parse_seqids']
        try:
            proc = subprocess.Popen(mkdb_arg, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out, error = proc.communicate()
            os.remove(mapped_seqfile)
        except OSError:
            logging.exception('Could not find makeblastdb. Make sure you include it in PATH!')
            raise OSError('Could not find makeblastdb. Make sure you include it in PATH!')
        
        if error:
            logging.exception('Error when calling makeblastdb to generate local database.\n' + error)
            raise OSError('Error when calling makeblastdb to generate local database.\n' + error)       

        # save the database size in txt file
        f1 = open(dbname + '.txt','w')
        f1.write('database size:'+'\t'+str(dbsize))
        f1.close()	
        
        
    #print 'Size of database is: ' + str(dbsize)
    
    return [dbsize, dbname]


#======================================================================================

# Blast queries against local database by calling blastp
#======================================================================================

# default parameters for blastp. Usually it works well.
task = 'blastp-short'		# Task to execute
#task = 'blastp'
matrix = 'PAM30'          # Scoring matrix name
comp_based_stats = 'F'		# No composition-based statistics
outfmt = '7 qaccver saccver nident length qlen mismatch gapopen gaps qstart qend sstart send qseq sseq evalue'			# alignment view options, 7 = Tabular with comment lines
word_size = '2'			# Word size for wordfinder algorithm
#gapopen = '32767'			# Cost to open a gap. 
#gapextend = '32767'			# Cost to extend a gap. 
# Allowable (gapopen, gapextend) pairs in blastp: (7,2), (6,2), (5,2), (10,1), (9,1), (8,1), (15,3), (14,2), (14,1), (13,3), (32767,32767)

# Search the database using blastp
# params includes blastp parameters dbsize, qsize, query, outfile, database, evalue, gapscore
def SeqBlast(params):
    
    # logging setup
    logging.basicConfig(filename=params.logfile, level=logging.INFO, format='%(message)s')

    # Calculate the search space
    searchsp = str(params.dbsize * params.qsize)
    
    # set maximum number of aligned sequences to keep in blastp search
    if params.dbsize < 10000:
        max_seq = params.dbsize
    else:
        max_seq = 10000

    # check output filename
    if params.outfile == '':
        output = params.query[:-6] + '_blast'
    else:
        output = params.outfile
    
    # run blast search according to the provided parameters
    blast_arg = ['blastp', '-task', task, '-matrix', matrix, '-db', params.database, '-query', params.query, '-out', output, '-evalue', str(params.evalue), '-comp_based_stats', comp_based_stats, 
'-dbsize', str(params.dbsize), '-searchsp', searchsp, '-word_size', word_size,'-gapopen', str(params.gapscore[0]), '-gapextend', str(params.gapscore[1]),  '-outfmt', outfmt, '-max_target_seqs', str(max_seq)]

    try:
        proc = subprocess.Popen(blast_arg)
        out, error = proc.communicate()
        
    except OSError:
        logging.exception('Could not find blastp. Make sure you include it in PATH!')
        raise OSError('Could not find blastp. Make sure you include it in PATH!')
        
    if error:
        logging.exception('Error when calling blastp to search local database.\n' + error)
        raise OSError('Error when calling blastp to search local database.\n' + error)



